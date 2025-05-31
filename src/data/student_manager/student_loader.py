import openpyxl
import logging
from config.config_manager import ConfigManager
from model.student import Student

def load_students(file_path):
    """加载学生数据（修复缩进和重复逻辑）"""
    try:
        config = ConfigManager.load_config()
        default_weight = config["weights"]["default"]
        custom_weights = config["weights"]["custom_weights"]
        
        # 加载Excel文件
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        # 检查表头是否包含权重列
        headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []
        has_weight = "权重" in headers
        
        students = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 2:  # 至少需要学号和姓名列
                continue
            student_id, name = row[0], row[1]
            # 优先使用表中权重，否则使用配置权重
            weight = row[2] if has_weight and len(row) > 2 else custom_weights.get(student_id, default_weight)
            students.append(Student(str(student_id), name, weight))
        return students
    except Exception as e:
        logging.error(f"加载学生名单失败: {str(e)}")
        raise