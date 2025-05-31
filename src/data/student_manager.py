import json  # 添加这行导入
import openpyxl
import os
import csv
import logging
import shutil
from datetime import datetime, time
from typing import List
from model.student import Student
from config.config_manager import ConfigManager  # 添加这行导入

class StudentManager:
    @staticmethod

    @staticmethod
    def _validate_format(file_path):
        """验证Excel文件格式"""
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []
            return headers[:2] == ["学号", "姓名"]  # 前两列必须为学号和姓名
        except Exception as e:
            logging.error(f"Excel文件验证失败: {str(e)}")
            return False

    @staticmethod
    def _create_template(file_path):
        """创建学生名单模板（含权重列）"""
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["学号", "姓名", "权重"])  # 标准表头
        sheet.append(["1001", "张三", None])
        sheet.append(["1002", "李四", None])
        wb.save(file_path)

    @staticmethod
    def load_students(file_path: str) -> List[Student]:
        """统一加载学生数据（合并重复方法）"""
        return student_loader.load_students(file_path)  # 调用student_loader的实现

    @staticmethod
    def save_students(file_path, students):
        """保存学生数据（保留权重信息）"""
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["学号", "姓名", "权重"])
        for student in students:
            sheet.append([student.id, student.name, student.weight])
        wb.save(file_path)

    @classmethod
    def backup_students(cls, file_path):
        """创建学生名单备份"""
        config = ConfigManager.load_config()
        backup_dir = os.path.join(config["backup_dir"], datetime.now().strftime("%Y%m%d"))
        os.makedirs(backup_dir, exist_ok=True)
        backup_file = os.path.join(backup_dir, f"students_{int(datetime.now().timestamp())}.xlsx")
        shutil.copyfile(file_path, backup_file)
        logging.info(f"学生名单备份成功: {backup_file}")

    @classmethod
    def add_student(cls, student_id, name):
        """添加学生（增加权重处理）"""
        try:
            config = ConfigManager.load_config()
            file_path = config["student_file"]
            cls.backup_students(file_path)
            
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            # 确保存在权重列（如果表头无"权重"则添加）
            headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []
            if "权重" not in headers:
                sheet.cell(row=1, column=3, value="权重")
            # 添加学生时默认使用配置中的默认权重
            sheet.append([student_id, name, config["weights"]["default"]])
            wb.save(file_path)
            logging.info(f"成功添加学生：{student_id} - {name}")
        except PermissionError as e:
            logging.error(f"文件访问失败：{str(e)}")
            raise
        except Exception as e:
            logging.error(f"添加学生失败：{str(e)}")
            raise